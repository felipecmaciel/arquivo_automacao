import tkinter as tk
from tkinter import filedialog, messagebox
from tkinter import ttk
from PyPDF2 import PdfMerger
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
import openpyxl
import time
import os

def login_inicial(driver):
    try:
        WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.ID, 'form:j_id_jsp_1714405282_5_body')))
        driver.find_element(By.NAME, 'form:j_id_jsp_1714405282_13').send_keys('teste')
        driver.find_element(By.NAME, 'form:senha').send_keys('teste')
        driver.find_element(By.NAME, 'form:btnOK').click()
        driver.implicitly_wait(3)
        driver.find_element(By.CLASS_NAME, 'rich-label-text-decor').click()
        driver.find_element(By.ID, 'formulario:formMenu:j_id_jsp_1387994506_5pc2:anchor').click()
        driver.find_element(By.ID, 'formulario:formMenu:j_id_jsp_1387994506_8pc2:anchor').click()
    except Exception as e:
        print(f'Login failed: {e}')

def repeticao_ait(driver, planilha, progress_var, total_rows):
    current_row = 0
    for linha in planilha.iter_rows(min_row=8):
        ait = linha[0].value
        try:
            aitbox = driver.find_element(By.NAME, 'formulario:j_id_jsp_1174046890_16')
            aitbox.clear()
            aitbox.send_keys(ait)
            driver.find_element(By.NAME, 'formulario:j_id_jsp_1174046890_65').click()
            driver.find_element(By.CLASS_NAME, 'imagem_sem_borda').click()
            driver.find_element(By.NAME, 'formulario:j_id_jsp_664462502_5').click()
            driver.implicitly_wait(5)
            driver.find_element(By.NAME, 'formulario:j_id_jsp_664462502_9').click()
            driver.find_element(By.ID, 'formulario:dados_shifted').click()
            driver.find_element(By.NAME, 'formulario:j_id_jsp_1174046890_64').click()
        except Exception as e:
            pass
        else:
            current_row += 1
            progress_var.set(current_row / total_rows * 100)
            progress_bar.update_idletasks()
        print(f'Error processing AIT {ait}: {e}')

def processa_planilha():
    file_path = filedialog.askopenfilename(filetypes=[('Excel Files', '*.xlsx')])
    if not file_path:
        return
    workbook = openpyxl.load_workbook(file_path)
    sheet_name = workbook.sheetnames[0]
    planilha = workbook[sheet_name]
    total_rows = len(list(planilha.iter_rows(min_row=8)))
    driver = webdriver.Chrome()
    driver.get('https://aite-gestao.pbh.gov.br/bh46web/')
    login_inicial(driver)
    repeticao_ait(driver, planilha, progress_var, total_rows)
    driver.quit()
    workbook.close()
    messagebox.showinfo('Processo Concluído', 'Os downloads foram concluídos com sucesso!')

def merge_pdfs():
    pdf_paths = filedialog.askopenfilenames(title='Selecione os PDFs para mesclar', filetypes=[('PDF Files', '*.pdf')])
    if not pdf_paths:
        return
    output_path = filedialog.asksaveasfilename(defaultextension='.pdf', filetypes=[('PDF Files', '*.pdf')])
    if not output_path:
        return
    merger = PdfMerger()
    try:
        for pdf in pdf_paths:
            merger.append(pdf)
        merger.write(output_path)
        merger.close()
        messagebox.showinfo('Sucesso', f'Os PDFs foram mesclados com sucesso em:\n{output_path}')
    except Exception as e:
        messagebox.showerror('Erro', f'Ocorreu um erro ao mesclar os PDFs: {e}')
        if os.path.exists(output_path):
            os.remove(output_path)
    else:
        pass

def main():
    global progress_var
    global progress_bar
    root = tk.Tk()
    root.title('Automatização de AITs e Mesclagem de PDFs')
    root.geometry('400x400')
    root.configure(bg='#f0f0f0')
    frame = tk.Frame(root, bg='#f0f0f0')
    frame.pack(padx=20, pady=20, fill=tk.BOTH, expand=True)
    title = tk.Label(frame, text='Selecione uma Opção', font=('Helvetica', 16, 'bold'), bg='#f0f0f0', fg='#333')
    title.pack(pady=(0, 10))
    button_planilha = tk.Button(frame, text='Processar Planilha', command=processa_planilha, font=('Helvetica', 12), bg='#4CAF50', fg='white', padx=10, pady=5)
    button_planilha.pack(pady=10)
    button_pdf = tk.Button(frame, text='Mesclar PDFs', command=merge_pdfs, font=('Helvetica', 12), bg='#2196F3', fg='white', padx=10, pady=5)
    button_pdf.pack(pady=10)
    progress_var = tk.DoubleVar()
    progress_bar = ttk.Progressbar(frame, variable=progress_var, maximum=100)
    progress_bar.pack(pady=20, fill=tk.X, padx=20)
    footer = tk.Label(frame, text='© 2024 - Felipe Maciel Corrêa', bg='#f0f0f0', fg='#777')
    footer.pack(side=tk.BOTTOM, pady=(10, 0))
    root.mainloop()
if __name__ == '__main__':
    main()